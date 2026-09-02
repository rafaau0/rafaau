from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


PSD_NAME = "lideranca.psd"
XLSX_NAME = "produtos.xlsx"
OUTPUT_DIR = "saida"
MAX_BOXES = 24


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def candidate_dirs() -> list[Path]:
    base = app_dir()
    dirs = [Path.cwd(), base, base.parent]
    unique: list[Path] = []
    for directory in dirs:
        if directory not in unique:
            unique.append(directory)
    return unique


def find_input_file(name: str) -> Path:
    for directory in candidate_dirs():
        path = directory / name
        if path.exists():
            return path
    searched = "\n".join(f"- {directory / name}" for directory in candidate_dirs())
    raise FileNotFoundError(f"Nao encontrei {name}. Procurei em:\n{searched}")


def find_photoshop() -> Path:
    known_paths = [
        Path(r"C:\Program Files\Adobe\Adobe Photoshop 2026\Photoshop.exe"),
        Path(r"C:\Program Files\Adobe\Adobe Photoshop 2025\Photoshop.exe"),
        Path(r"C:\Program Files\Adobe\Adobe Photoshop 2024\Photoshop.exe"),
    ]
    for path in known_paths:
        if path.exists():
            return path

    adobe_dir = Path(r"C:\Program Files\Adobe")
    if adobe_dir.exists():
        matches = sorted(adobe_dir.glob("Adobe Photoshop */Photoshop.exe"), reverse=True)
        if matches:
            return matches[0]

    raise FileNotFoundError("Nao encontrei o Photoshop instalado em C:\\Program Files\\Adobe.")


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    replacements = {
        "Ç": "C",
        "Ã": "A",
        "Á": "A",
        "À": "A",
        "Â": "A",
        "É": "E",
        "Ê": "E",
        "Í": "I",
        "Ó": "O",
        "Ô": "O",
        "Õ": "O",
        "Ú": "U",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text


def format_price(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}".replace(".", ",")
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace("R$", "").strip()
    return text


def load_products(xlsx_path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    headers = [normalize_header(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    desc_col = next((i + 1 for i, header in enumerate(headers) if "DESCR" in header), 1)
    price_col = next((i + 1 for i, header in enumerate(headers) if header in {"VALOR", "PRECO", "PRECO R$"} or "VALOR" in header or "PRECO" in header), 2)

    products: list[dict[str, str]] = []
    for row in range(2, sheet.max_row + 1):
        description = str(sheet.cell(row, desc_col).value or "").strip()
        price = format_price(sheet.cell(row, price_col).value)
        if not description and not price:
            continue
        products.append({"description": description, "price": price})
        if len(products) >= MAX_BOXES:
            break

    if not products:
        raise ValueError("A planilha nao tem produtos preenchidos.")
    return products


def js_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def build_jsx(psd_path: Path, output_path: Path, products: list[dict[str, str]], data_text: str) -> str:
    products_json = json.dumps(products, ensure_ascii=False, indent=2)
    return f"""#target photoshop
app.displayDialogs = DialogModes.NO;

var psdFile = new File({js_string(str(psd_path))});
var outputFile = new File({js_string(str(output_path))});
var products = {products_json};
var dataText = {js_string(data_text)};

function pad2(number) {{
  return number < 10 ? "0" + number : String(number);
}}

function layerName(layer) {{
  try {{ return String(layer.name); }} catch (error) {{ return ""; }}
}}

function findLayerSet(container, targetName) {{
  for (var i = 0; i < container.layers.length; i++) {{
    var layer = container.layers[i];
    if (layer.typename === "LayerSet") {{
      if (layerName(layer) === targetName) {{
        return layer;
      }}
      var nested = findLayerSet(layer, targetName);
      if (nested) {{
        return nested;
      }}
    }}
  }}
  return null;
}}

function findArtLayerByNames(container, names) {{
  for (var i = 0; i < container.layers.length; i++) {{
    var layer = container.layers[i];
    var currentName = layerName(layer);
    for (var n = 0; n < names.length; n++) {{
      if (currentName === names[n]) {{
        return layer;
      }}
    }}
    if (layer.typename === "LayerSet") {{
      var nested = findArtLayerByNames(layer, names);
      if (nested) {{
        return nested;
      }}
    }}
  }}
  return null;
}}

function findAllArtLayersByNames(container, names, found) {{
  found = found || [];
  for (var i = 0; i < container.layers.length; i++) {{
    var layer = container.layers[i];
    var currentName = layerName(layer);
    for (var n = 0; n < names.length; n++) {{
      if (currentName === names[n]) {{
        found.push(layer);
        break;
      }}
    }}
    if (layer.typename === "LayerSet") {{
      findAllArtLayersByNames(layer, names, found);
    }}
  }}
  return found;
}}

function replaceOfferDate(text, dateValue) {{
  if (!dateValue) {{
    return text;
  }}

  var lines = text.split(/(\\r\\n|\\r|\\n)/);
  for (var lineWithDate = 0; lineWithDate < lines.length; lineWithDate += 2) {{
    if (lines[lineWithDate].indexOf("/") >= 0) {{
      var existingLine = lines[lineWithDate];
      var existingComma = existingLine.indexOf(",");
      lines[lineWithDate] = existingComma >= 0 ? dateValue + existingLine.substring(existingComma) : dateValue;
      return lines.join("");
    }}
  }}

  var patterns = [
    /[^,\\r\\n]*[\\/][^,\\r\\n]*/i,
    /\\d{{1,2}}\\s*A\\s*\\d{{1,2}}\\/\\d{{1,2}}\\/\\d{{2,4}}/i,
    /\\d{{1,2}}\\/\\d{{1,2}}\\/\\d{{2,4}}\\s*A\\s*\\d{{1,2}}\\/\\d{{1,2}}\\/\\d{{2,4}}/i,
    /\\d{{1,2}}\\/\\d{{1,2}}\\/\\d{{2,4}}/i
  ];

  for (var i = 0; i < patterns.length; i++) {{
    if (patterns[i].test(text)) {{
      return text.replace(patterns[i], dateValue);
    }}
  }}

  for (var lineIndex = 0; lineIndex < lines.length; lineIndex += 2) {{
    if (/VALID|VÁLID/i.test(lines[lineIndex]) && lineIndex + 2 < lines.length) {{
      var nextLine = lines[lineIndex + 2];
      var commaIndex = nextLine.indexOf(",");
      lines[lineIndex + 2] = commaIndex >= 0 ? dateValue + nextLine.substring(commaIndex) : dateValue;
      return lines.join("");
    }}
  }}

  return text;
}}

function setText(layer, value) {{
  if (!layer) {{
    return false;
  }}
  if (layer.typename !== "ArtLayer" || layer.kind !== LayerKind.TEXT) {{
    return false;
  }}
  var textItem = layer.textItem;
  var style = {{}};
  var keys = [
    "font",
    "size",
    "color",
    "justification",
    "position",
    "leading",
    "tracking",
    "horizontalScale",
    "verticalScale",
    "baselineShift",
    "fauxBold",
    "fauxItalic",
    "allCaps",
    "underline"
  ];

  for (var i = 0; i < keys.length; i++) {{
    try {{
      style[keys[i]] = textItem[keys[i]];
    }} catch (error) {{}}
  }}

  textItem.contents = value;

  for (var j = 0; j < keys.length; j++) {{
    try {{
      if (style[keys[j]] !== undefined) {{
        textItem[keys[j]] = style[keys[j]];
      }}
    }} catch (error) {{}}
  }}
  return true;
}}

function cTID(value) {{
  return charIDToTypeID(value);
}}

function sTID(value) {{
  return stringIDToTypeID(value);
}}

function getTextLayerDescriptor(layer) {{
  var reference = new ActionReference();
  reference.putIdentifier(cTID("Lyr "), layer.id);
  return executeActionGet(reference).getObjectValue(sTID("textKey"));
}}

function applyTextLayerDescriptor(layer, textDescriptor) {{
  var reference = new ActionReference();
  reference.putIdentifier(cTID("Lyr "), layer.id);

  var descriptor = new ActionDescriptor();
  descriptor.putReference(cTID("null"), reference);
  descriptor.putObject(cTID("T   "), sTID("textLayer"), textDescriptor);
  executeAction(cTID("setd"), descriptor, DialogModes.NO);
}}

function getStyleSize(styleDescriptor) {{
  try {{
    return styleDescriptor.getUnitDoubleValue(sTID("size"));
  }} catch (error) {{
    return 0;
  }}
}}

function copyFirstParagraphStyleRange(textDescriptor, textLength) {{
  var output = new ActionList();
  try {{
    var existing = textDescriptor.getList(sTID("paragraphStyleRange"));
    if (existing.count > 0) {{
      var first = existing.getObjectValue(0);
      var item = new ActionDescriptor();
      item.putInteger(sTID("from"), 0);
      item.putInteger(sTID("to"), textLength);
      item.putObject(sTID("paragraphStyle"), sTID("paragraphStyle"), first.getObjectValue(sTID("paragraphStyle")));
      output.putObject(sTID("paragraphStyleRange"), item);
      return output;
    }}
  }} catch (error) {{}}
  return null;
}}

function setPriceText(layer, value) {{
  if (!layer) {{
    return false;
  }}
  if (layer.typename !== "ArtLayer" || layer.kind !== LayerKind.TEXT) {{
    return false;
  }}

  try {{
    var textDescriptor = getTextLayerDescriptor(layer);
    var styleRanges = textDescriptor.getList(sTID("textStyleRange"));
    if (styleRanges.count < 2) {{
      return setText(layer, value);
    }}

    var largestStyle = null;
    var smallestStyle = null;
    var largestSize = -1;
    var smallestSize = 999999;

    for (var i = 0; i < styleRanges.count; i++) {{
      var styleRange = styleRanges.getObjectValue(i);
      var textStyle = styleRange.getObjectValue(sTID("textStyle"));
      var size = getStyleSize(textStyle);
      if (size > largestSize) {{
        largestSize = size;
        largestStyle = textStyle;
      }}
      if (size > 0 && size < smallestSize) {{
        smallestSize = size;
        smallestStyle = textStyle;
      }}
    }}

    if (!largestStyle || !smallestStyle) {{
      return setText(layer, value);
    }}

    var commaIndex = value.indexOf(",");
    var integerEnd = commaIndex > 0 ? commaIndex : value.length;
    var newStyleRanges = new ActionList();

    var integerRange = new ActionDescriptor();
    integerRange.putInteger(sTID("from"), 0);
    integerRange.putInteger(sTID("to"), integerEnd);
    integerRange.putObject(sTID("textStyle"), sTID("textStyle"), largestStyle);
    newStyleRanges.putObject(sTID("textStyleRange"), integerRange);

    if (integerEnd < value.length) {{
      var centsRange = new ActionDescriptor();
      centsRange.putInteger(sTID("from"), integerEnd);
      centsRange.putInteger(sTID("to"), value.length);
      centsRange.putObject(sTID("textStyle"), sTID("textStyle"), smallestStyle);
      newStyleRanges.putObject(sTID("textStyleRange"), centsRange);
    }}

    textDescriptor.putString(cTID("Txt "), value);
    textDescriptor.putList(sTID("textStyleRange"), newStyleRanges);

    var paragraphRanges = copyFirstParagraphStyleRange(textDescriptor, value.length);
    if (paragraphRanges) {{
      textDescriptor.putList(sTID("paragraphStyleRange"), paragraphRanges);
    }}

    applyTextLayerDescriptor(layer, textDescriptor);
    return true;
  }} catch (error) {{
    return setText(layer, value);
  }}
}}

var doc = app.open(psdFile);
var changed = 0;
var missing = [];

if (dataText) {{
  var dataLayers = findAllArtLayersByNames(doc, ["DATA_01"]);
  if (dataLayers.length) {{
    for (var dl = 0; dl < dataLayers.length; dl++) {{
      if (dataLayers[dl] && dataLayers[dl].typename === "ArtLayer" && dataLayers[dl].kind === LayerKind.TEXT) {{
        if (setText(dataLayers[dl], dataText)) {{
          changed++;
        }}
      }}
    }}
  }} else {{
    missing.push("DATA_01");
  }}
}}

for (var i = 0; i < products.length; i++) {{
  var number = pad2(i + 1);
  var group = findLayerSet(doc, "BOX_" + number);
  if (!group) {{
    missing.push("BOX_" + number);
    continue;
  }}

  var descriptionLayer = findArtLayerByNames(group, [
    "DESCRICAO_" + number,
    "DESCRIÇÃO_" + number,
    "DESCRIÇAO_" + number
  ]);
  var priceLayer = findArtLayerByNames(group, [
    "VALOR_" + number,
    "PRECO_" + number,
    "PREÇO_" + number
  ]);

  if (setText(descriptionLayer, products[i].description)) {{
    changed++;
  }} else {{
    missing.push("DESCRICAO_" + number);
  }}

  if (setPriceText(priceLayer, products[i].price)) {{
    changed++;
  }} else {{
    missing.push("PRECO_" + number + " / VALOR_" + number);
  }}
}}

doc.saveAs(outputFile);
doc.close(SaveOptions.DONOTSAVECHANGES);

var logFile = new File(outputFile.fsName.replace(/\\.psd$/i, "_log.txt"));
logFile.encoding = "UTF-8";
logFile.open("w");
logFile.writeln("Arquivo gerado: " + outputFile.fsName);
logFile.writeln("Produtos lidos: " + products.length);
if (dataText) {{
  logFile.writeln("Texto DATA_01 informado: " + dataText);
}}
logFile.writeln("Textos alterados: " + changed);
if (missing.length) {{
  logFile.writeln("");
  logFile.writeln("Itens nao encontrados:");
  for (var m = 0; m < missing.length; m++) {{
    logFile.writeln("- " + missing[m]);
  }}
}}
logFile.close();
"""


def run_photoshop_script(photoshop_path: Path, jsx_path: Path) -> None:
    vbs_path = Path(tempfile.gettempdir()) / "preencher_psd_photoshop.vbs"
    vbs = f"""Set app = CreateObject("Photoshop.Application")
app.Visible = True
app.DoJavaScriptFile "{str(jsx_path)}"
"""
    vbs_path.write_text(vbs, encoding="utf-16")
    subprocess.run(["cscript.exe", "//nologo", str(vbs_path)], check=True)


def main() -> int:
    try:
        psd_path = find_input_file(PSD_NAME)
        xlsx_path = find_input_file(XLSX_NAME)
        photoshop_path = find_photoshop()

        products = load_products(xlsx_path)
        data_text = input("Digite o texto para DATA_01 (use \\n para quebrar linha). Enter para nao alterar: ").strip()
        data_text = data_text.replace("\\n", "\r")
        output_dir = psd_path.parent / OUTPUT_DIR
        output_dir.mkdir(exist_ok=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"lideranca_preenchido_{stamp}.psd"
        jsx_path = Path(tempfile.gettempdir()) / "preencher_psd.jsx"
        jsx_path.write_text(build_jsx(psd_path, output_path, products, data_text), encoding="utf-8-sig")

        print(f"PSD: {psd_path}")
        print(f"Planilha: {xlsx_path}")
        print(f"Produtos: {len(products)}")
        if data_text:
            print("Texto DATA_01 informado.")
        print("Abrindo Photoshop para preencher o arquivo...")

        run_photoshop_script(photoshop_path, jsx_path)

        print("")
        print("Concluido.")
        print(f"Arquivo gerado: {output_path}")
        print(f"Log: {output_path.with_name(output_path.stem + '_log.txt')}")
        return 0
    except Exception as error:
        print("")
        print("Erro ao preencher o PSD:")
        print(error)
        input("Pressione Enter para fechar...")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

